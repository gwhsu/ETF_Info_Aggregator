import tkinter as tk
from tkinter import messagebox
import pandas as pd
import openpyxl


class DividendYieldCalculator:
    def __init__(self, master):
        self.master = master
        self.master.title("Dividend Yield Calculator")

        # Create frame for input fields
        self.frame = tk.Frame(self.master)
        self.frame.pack(padx=10, pady=10)

        # Initialize lists to store entry widgets for ETF names and asset allocation percentages
        self.entry_etf_names = []
        self.entry_allocation_percentages = []

        # Add entry fields for ETF names and asset allocation percentages
        for i in range(10):
            label_etf_name = tk.Label(self.frame, text=f"ETF Name {i+1}:")
            label_etf_name.grid(row=i, column=0, padx=5, pady=5, sticky="w")
            entry_etf_name = tk.Entry(self.frame, width=30)
            entry_etf_name.grid(row=i, column=1, padx=5, pady=5)
            self.entry_etf_names.append(entry_etf_name)

            label_allocation_percentage = tk.Label(self.frame, text=f"Allocation Percentage {i+1}:")
            label_allocation_percentage.grid(row=i, column=2, padx=5, pady=5, sticky="w")
            entry_allocation_percentage = tk.Entry(self.frame, width=10)
            entry_allocation_percentage.grid(row=i, column=3, padx=5, pady=5)
            self.entry_allocation_percentages.append(entry_allocation_percentage)

        # Add save button
        self.save_button = tk.Button(self.master, text="Save Selected Stocks", command=self.save_selected_stocks)
        self.save_button.pack(pady=5)

    def calculate_div_yield(self):
        # 建立一個新的工作簿
        wb = openpyxl.Workbook()

        # 選擇活動工作表
        ws = wb.active

        # # 在 A1 儲存格中添加一個超連結
        # ws['A1'] = "點我"
        # ws['A1'].hyperlink = "http://www.example.com"

        # # 儲存工作簿
        # wb.save('hyperlink_example.xlsx')

        # Read the Excel file
        df = pd.read_excel("2023_all.xlsx")

        # Initialize lists to store ETF names, asset allocation percentages, and selected stock codes
        etf_names = []
        allocation_percentages = []
        selected_stock_codes = []

        # Get user input for ETF names and asset allocation percentages
        for entry_etf_name, entry_allocation_percentage in zip(self.entry_etf_names, self.entry_allocation_percentages):
            etf_name = entry_etf_name.get().strip()
            allocation_percentage = entry_allocation_percentage.get().strip()
            if etf_name and allocation_percentage:
                etf_names.append(etf_name)
                allocation_percentages.append(float(allocation_percentage))
                selected_stock_codes.append(etf_name)
        print(selected_stock_codes)
        print(allocation_percentages)
        # # Select rows with specific stock codes
        # selected_df = df[df['code'].isin(selected_stock_codes)].copy()
        #
        # # Add allocation percentage column
        # selected_df['Allocation Percentage'] = allocation_percentages
        # Create a DataFrame to store the user input
        user_input_df = pd.DataFrame({'ETF Name': etf_names, 'Allocation Percentage': allocation_percentages})

        # Merge the user input DataFrame with the original DataFrame based on ETF Name
        selected_df = pd.merge(user_input_df, df, how='inner', left_on='ETF Name', right_on='code')

        # Drop the ETF Name column
        selected_df.drop(columns=['ETF Name'], inplace=True)
        print(selected_df)

        # Calculate the total investment amount
        total_investment = sum(selected_df['Price'] * selected_df['Allocation Percentage'] / 100)

        # Calculate the weighted average dividend yield
        weighted_avg_div_yield = (selected_df['Price'] * selected_df['殖利率'] * selected_df[
            'Allocation Percentage'] / 100).sum() / total_investment

        # Calculate overall dividend yield based on asset allocation
        overall_div_yield = sum(
            [allocation * div_yield for allocation, div_yield in zip(allocation_percentages, selected_df['殖利率'])]) / 100

        # Calculate the total return (dividend yield + price appreciation/depreciation)
        total_return = overall_div_yield
        total_Percentage = selected_df['Allocation Percentage'].sum()
        # Append a row for the weighted average dividend yield
        overall_row = pd.DataFrame({'code': ['Overall'], 'Price': [None], '殖利率': [overall_div_yield], 'Allocation Percentage': [total_Percentage]})
        selected_df = pd.concat([selected_df, overall_row], ignore_index=True)

        # Show message box with the total return
        messagebox.showinfo("Total Return", f"The total return (dividend yield) is: {total_return:.2f}%")

        # Store the selected stocks DataFrame
        self.selected_df = selected_df

    def save_selected_stocks(self):
        # Automatically calculate dividend yield before saving
        self.calculate_div_yield()

        try:
            # Save the selected stocks to a new Excel file
            self.selected_df.to_excel("selected_stocks.xlsx", index=False)
            messagebox.showinfo("Success", "Selected stocks saved to 'selected_stocks.xlsx'.")
        except AttributeError:
            messagebox.showerror("Error", "No selected stocks to save. Please enter ETF names and asset allocation percentages.")

def main():
    root = tk.Tk()
    app = DividendYieldCalculator(root)
    root.mainloop()

if __name__ == "__main__":
    main()
